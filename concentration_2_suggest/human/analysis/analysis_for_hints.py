import os
import pandas as pd

from openpyxl import load_workbook

ids = []
mod = []
turns = []
time_hints = []
type_hints = []
followed = []
wrong = []
match = []

# get all csv 
path ="..\data\\"
df_append = pd.DataFrame()
for root, dirs, files in sorted(os.walk(path)):
    for file in files:
        if(file.endswith(".csv")):
            print(os.path.join(root,file))
            # one data frame for file
            df = pd.read_csv(os.path.join(root,file), sep=';')
            print(df)

            # Filtra le righe in cui suggestion_type non è "none"
            filtered_df = df[df['suggestion_type'] != 'none']
            
            # Aggiungi le informazioni alle liste
            ids.extend(filtered_df['id_player'])
            mod.extend(filtered_df['experiment_condition'])
            turns.extend(filtered_df['turn_number'])
            time_hints.extend(filtered_df['time_game'])
            type_hints.extend(filtered_df['suggestion_type'])
            wrong.extend(filtered_df['wrong_hint'])

            for suggestion_type, suggested, clicked in zip(filtered_df['suggestion_type'], filtered_df['position_suggested'], filtered_df['position_clicked']):
                clicked = eval(clicked)
                suggested = eval(suggested)

                if suggestion_type == 'card':
                    followed.append(clicked == suggested)
                elif suggestion_type == 'row':
                    # Verifica se la riga suggerita è uguale alla riga cliccata
                    print(suggestion_type, clicked, "row: ", clicked[0], "sugg: ", suggested, "followed: ", clicked[0] == suggested)
                    followed.append(clicked[0] == suggested)
                elif suggestion_type == 'column':
                    # Verifica se la colonna suggerita è uguale alla colonna cliccata
                    print(suggestion_type, clicked, "col: ", clicked[1], "sugg: ", suggested, "followed: ", clicked[1] == suggested)
                    followed.append(clicked[1] == suggested)
                else:
                    followed.append(False)
            
            match.extend(filtered_df['match'])

csv_struct = {
    'id_player': ids,
    'experiment_condition': mod,
    'turn_number': turns,
    'time_hint': time_hints,
    'hint_type': type_hints,
    'wrong_hint': wrong,
    'hint_followed': followed,
    'match': match
}

# create dataframe
df_deception = pd.DataFrame({key: [value[i] for i, robot_type in enumerate(mod) if robot_type == 'deception'] for key, value in csv_struct.items()})

# id_player must be int
df_deception['id_player'] = df_deception['id_player'].astype(int)

# sort df 
df_deception = df_deception.sort_values(by=['id_player', 'turn_number'])

# handle time
def fix_time_format(time_str):
    # crea la stringa in hh:mm:ss
    parts = time_str.split(':')
    if len(parts) == 2:  # Se manca l'ora
        hours = 0
        minutes, seconds = map(int, parts)
    elif len(parts) == 3:  # Se sono presenti ore, minuti e secondi
        hours, minutes, seconds = map(int, parts)
    else:
        return None  # Formato non valido

    # aggiunge gli zeri mancanti
    formatted_time = '{:02d}:{:02d}:{:02d}'.format(hours, minutes, seconds)
    return formatted_time

dataframes = [df_deception]
for df in dataframes:
    # Applica la funzione di correzione alla colonna 'time to finish'
    df['time_hint'] = df['time_hint'].apply(fix_time_format)

# get the existing file and remove sheets of deceiving modes if available
wb = load_workbook("results.xlsx")
if "Deception" in wb.sheetnames:
    wb.remove(wb["Deception"])
wb.save("results.xlsx")

# open results and write the other sheets
with pd.ExcelWriter("results.xlsx", mode="a") as writer:
    # Aggiunta dei DataFrame come nuovi fogli nel file Excel
    df_deception.to_excel(writer, sheet_name="Deception", index=False)
